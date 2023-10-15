#!/usr/bin/env python3

import argparse
import json
import xlsxwriter
import openpyxl
import os

from datetime import datetime, timezone
import pytz


def parse_arguments():
    parser = argparse.ArgumentParser(description="Read Favourite column from one XLSX file and add values to the generated XLSX file by a common column (ID).  It'll also show which catalog items are new and any changes to AWS Favorites.  Favorites that no longer exist will be logged to the console.")
    parser.add_argument("source_file", nargs='?', default="", help="Path to the source XLSX file (optional)")
    return parser.parse_args()


def to_excel_date( d ):
    if d != None:
        return d.strftime('%x %X')
    else:
        return None


def add_item( dfs, type, session ):

    category = dfs.get(type, None)
    if category is None:
        category = []
        dfs[type] = category

    item = {}
    item["selected"] = session['selected']
    item["Favourite"] = session['favourite']
    item["FavoriteAWS"] = session['isFavorite']
    item["IsNew"] = session['isNew']
    item["ID"] = session['thirdPartyID']
    item["Title"] = session['title']
    item["SessionLevel"] = session['sessionLevel']
    item["Description"] = session['description']
    item["Type"] = session['sessionType']
    item["TrackName"] = session['trackName']
    item["Venue"] = session['venue']
    item["Day"] = session['day']
    item["StartTime"] = to_excel_date(session['startTime'])
    item["EndTime"] = to_excel_date(session['endTime'])
    item["scheduleUid"] = session['scheduleUid']
    item["sessionUid"] = session['sessionUid']
    
    simple_tag_name = ""
    for t in session['tags']:
        simple_tag_name = simple_tag_name + t['parentTagName'].replace(" ", "_").lower() + ": " + t['tagName'] + ", "
    item["Tags"] = simple_tag_name
    
    category.append(item)

    
def parse_sessions(sessions, favourites, is_selected_data):
    items = {}
    print( f"There are currently {len(sessions)} sessions" )

    # Display the sessionType, trackName, thirdPartyID, title, and description for each session
    for session in sessions:

        type = session['sessionType']
        
        # Update Favourites
        dest_id = session['thirdPartyID'].split("-")[0]
        session["favourite"] = favourites[dest_id] if dest_id in favourites else ""

        # Update selected
        dest_id = session['thirdPartyID']
        session["selected"] = is_selected_data[dest_id] if dest_id in is_selected_data else ""

        # Update IsNew
        session["isNew"] = False if dest_id in is_selected_data else True

        # Clean up date/times
        startTime = session['startDateTime']
        endTime = session['endDateTime']
        if startTime != "":
            startTime = datetime.fromtimestamp(startTime, timezone.utc).astimezone(pytz.timezone('America/Los_Angeles'))
        else:
            startTime = None
        if endTime != "":
            endTime = datetime.fromtimestamp(endTime, timezone.utc).astimezone(pytz.timezone('America/Los_Angeles'))
        else:
            endTime = None

        session["startTime"] = startTime
        session["endTime"] = endTime

        # Extract session level
        if len(session['thirdPartyID']) >= 4:
            session_level = int(session['thirdPartyID'][3])
        else:
            session_level = 0
        session["sessionLevel"] = session_level

        # Get venue and date from tags
        # each tag has a parentTagName of either Venue or Day
        # and we want the tagName
        venue = ""
        day = ""
        for tag in session['tags']:
            if tag['parentTagName'] == "Venue":
                venue = tag['tagName']
            if tag['parentTagName'] == "Day":
                day = tag['tagName']

        session["venue"] = venue
        session["day"] = day

        add_item( items, "ALL", session )
    return items


def read_excel_source(path: str):
    favourites_data = {}
    is_selected_data = {}
    
    if os.path.exists(path):
        source_workbook = openpyxl.load_workbook(path)
        source_sheet = source_workbook['ALL']
        
        for row in source_sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers are in row 1
            favourite_value = row[1]  # Assuming the "Favourite" column is the second column (0-based index)
            isselected_value = row[0]  # Assuming the "Selected" column is the first column (0-based index)
            id_value = row[4].split('-')[0]  # Assuming the "ID" column is the second column (0-based index)
            if not id_value in favourites_data: 
                favourites_data[id_value] = favourite_value
            if not id_value in is_selected_data: 
                is_selected_data[id_value] = isselected_value
            
        source_workbook.close()
    
    return favourites_data, is_selected_data


def write_excel_destination(path: str, items):

    workbook = xlsxwriter.Workbook(path, {'default_row_height': 20})

    workbook = xlsxwriter.Workbook("reinvent.xlsx", {'default_row_height': 20})
    
    cell_format = workbook.add_format({'text_wrap': True, "font_size": 12})
    favorite_format = workbook.add_format({'text_wrap': False, "font_size": 12})
    bold_format = workbook.add_format({'bold': 1, 'text_wrap': True, "font_size": 12})

    columns = [
        ("selected", 5),
        ("Favourite", 5),
        ("FavoriteAWS", 8),
        ("IsNew", 7),
        ("ID", 12),
        ("Title", 80),
        ("Description", 110),
        ("TrackName", 16),
        ("Venue", 14),
        ("StartTime", 18),
        ("EndTime", 18),
        ("Day", 11),
        ("Type", 16),
        ("Tags", 110),
        ("scheduleUid", 20),
        ("sessionUid", 20),
    ]
    # Get the list of keys and move the ALL and Favourites to the top
    keys = list(items.keys())
    
    if "ALL" in keys: 
        keys.remove("ALL")
        keys.insert(0, "ALL")
        
    for key in keys:
        category = items.get(key, [])
        worksheet = workbook.add_worksheet(key)

        # Write the header row
        for col_num, (header, width) in enumerate(columns):
            worksheet.write(0, col_num, header, bold_format)
            worksheet.set_column(col_num, col_num, width)

        row = 1
        for item in category:
            element = item["Favourite"]
            if element is None or (isinstance(element, str) and element.strip() == ""):
                format = cell_format 
            else:
                format = favorite_format 


            for col_num, (column_name, _) in enumerate(columns):
                worksheet.write(row, col_num, item.get(column_name, ""), format)

            row += 1

    workbook.close()


if __name__ == "__main__":
    # Parse the command-line arguments
    args = parse_arguments()
    
    # If the user decides to load a previous xlsx generated via this tool then set path to it
    source_file_path = args.source_file
    
    # Load our sessions.json
    with open("sessions.json", "r") as f:
        sessions = json.load(f)
    
    # Read our source excel which now has our favourites (where we can whatever markup we want) 
    #    and another column where I keep track of those I want to select for realz
    favourites_data = []
    is_selected_data = []
    # Note: For the old xlsx file, be sure to remove filters and unfreeze rows else it'll break
    #       Continues to expect the ALL tab
    favourites_data, is_selected_data = read_excel_source(source_file_path)
    
    # Iterate thru Session data and do stuff to it for easy writing to Excel in next step
    #  - Cleanup date/times
    #  - extract different fields from tags and make them main elements in the dataset
    #  - apply selected element to dataset
    #  - Mark new items as new (since last excel spreadsheet)
    #  - apply Favourite element to dataset from excel
    items = parse_sessions(sessions, favourites_data, is_selected_data)
    
    # Write modified session data to Excel
    write_excel_destination("reinvent.xlsx", items)
